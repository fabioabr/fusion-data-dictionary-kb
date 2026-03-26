#!/usr/bin/env python3
"""
extract_pvos.py — Extrai dados de PVOs do BICC Mapping Excel e gera JSONs

Uso:
    python extract_pvos.py --module GL
    python extract_pvos.py --all
"""

import argparse
import json
import os
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)


MODULE_HINTS = {
    'GL': ['Gl', 'GL', 'FinGl', 'Ledger', 'Journal', 'Balance', 'Calendar'],
    'AP': ['Ap', 'AP', 'FinAp', 'Payable', 'Invoice', 'Payment', 'Check', 'Expense'],
    'AR': ['Ar', 'AR', 'FinAr', 'Receivable', 'Receipt', 'Collection', 'Revenue'],
    'PO': ['Po', 'PO', 'Procurement', 'Prc', 'Purchasing', 'Sourcing', 'Supplier', 'Requisition'],
    'HCM': [
        'Hcm', 'HCM', 'Person', 'Assignment', 'Compensation', 'Benefits', 'Benefit',
        'Payroll', 'Pay', 'Absence', 'Talent', 'Performance', 'Goal', 'Profile',
        'Learning', 'Workforce', 'Worker', 'Content', 'HealthandSafety',
        'TimeRepository', 'Organization', 'Concepts', 'Checklist',
        'Succession', 'Career', 'Mentor', 'Questionnaire',
    ],
}

# AM names that map to known modules (for paths not caught by MODULE_HINTS)
AM_MODULE_MAP = {
    # HCM
    'HcmProfileCoreAM': 'HCM', 'HCMExtractAM': 'HCM', 'PersonAM': 'HCM',
    'TimeRepositoryAM': 'HCM', 'OrganizationAM': 'HCM', 'ConceptsAnalyticsAM': 'HCM',
    'HcmGoalCoreAM': 'HCM', 'HealthandSafetyAM': 'HCM', 'AnalyticsServiceAM': 'HCM',
    'HcmBenefitsAM': 'HCM', 'CompensationAM': 'HCM', 'ContentAM': 'HCM',
    'HcmTalentCalibMeetingsAM': 'HCM', 'HcmPerformanceDocsAM': 'HCM',
    'HcmProfileContentLibraryAM': 'HCM', 'AssignmentAM': 'HCM',
    'PerformanceManagementAM': 'HCM', 'ChecklistAM': 'HCM',
    'WorkforceReputationAM': 'HCM', 'PersonDocumentDeliveryAM': 'HCM',
    'TalentAM': 'HCM', 'AbsenceAM': 'HCM', 'PayrollAM': 'HCM',
    'WellbeingAM': 'HCM', 'CareerDevAM': 'HCM',
    # SCM (Supply Chain Management)
    'ScmExtractAM': 'SCM', 'InventoryAM': 'SCM', 'WorkOrderAM': 'SCM',
    'EgpItemsPublicModelAM': 'SCM', 'DooTopAM': 'SCM',
    # FIN (generic Finance / shared)
    'FinExtractAM': 'FIN', 'FinFunBusinessUnitsAM': 'FIN',
    # PRJ (Projects)
    'PrjExtractAM': 'PRJ',
    # GRC (Governance Risk Compliance)
    'GRCTOPBIAM': 'GRC',
    # OKC (Contracts)
    'ContractsCoreAM': 'OKC',
    # GMS (Grants)
    'GmsSetupAM': 'GMS', 'GmsAwardAM': 'GMS',
    # PSC (Public Sector)
    'PscPermitsReportingAM': 'PSC',
    # STU (Student)
    'StudentEnrollmentAM': 'STU',
    # PRJ (Projects) — additional
    'PjfSetupTransactionsAM': 'PRJ', 'PjrResourceRequestAM': 'PRJ',
    'PjfProjectDefinitionAM': 'PRJ', 'PjfSetupProjectsAM': 'PRJ',
    'PjoBudgetsAndForecastsAnalyticsAM': 'PRJ', 'PjcTransactionsAM': 'PRJ',
    'PjfProjectAM': 'PRJ', 'PjcCapitalAM': 'PRJ', 'PjrReportingAM': 'PRJ',
    'PjeCommonAM': 'PRJ', 'PjeIssuesAM': 'PRJ', 'PjfSetupSystemAM': 'PRJ',
    'PjtProjectPlanAM': 'PRJ', 'PjfResourcesAnalyticsAM': 'PRJ',
    'PjfSetupStatusesAnalyticsAM': 'PRJ', 'PjlTaskManagementAM': 'PRJ',
    'PjrResourceAM': 'PRJ', 'PjrSetupAM': 'PRJ', 'PjtCommonAM': 'PRJ',
    'PjtPrjResourceAM': 'PRJ', 'PjtResourcesAM': 'PRJ', 'PjtSetupAM': 'PRJ',
    'PjfSetupPeriodExceptionsAM': 'PRJ', 'PjfSetupPeriodsAM': 'PRJ',
    'PjoFinancialPlanOptionsAnalyticsAM': 'PRJ', 'ContractProjectLinkAM': 'PRJ',
    # SCM (Supply Chain) — additional
    'MaintProgramAM': 'SCM', 'MscAnalyticsTopAM': 'SCM',
    'ChangeTypesAM': 'SCM', 'EgiBatchesPublicModelAnalyticsAM': 'SCM',
    'CatalogAM': 'SCM', 'EgpStructuresPublicModelAM': 'SCM',
    'ProductGenealogyAM': 'SCM', 'ChangeObjectsAM': 'SCM',
    'QualityActionsAnalyticsAM': 'SCM', 'RcvDimensionsAM': 'SCM',
    'InvOrgPublicViewAM': 'SCM', 'InvUomPublicViewAM': 'SCM',
    'RequirementsAnalyticsAM': 'SCM', 'IdeasAnalyticsAM': 'SCM',
    'OnhandValuationAM': 'SCM', 'QualityCommonAnalyticsAM': 'SCM',
    'SupplyCollaborationAM': 'SCM', 'ValuationUnitAM': 'SCM',
    'CostTransactionAM': 'SCM', 'InspectionPlansAM': 'SCM',
    'ItemClassesAM': 'SCM', 'PartsAnalyticsAM': 'SCM',
    'QualityIssuesAnalyticsAM': 'SCM', 'WorkCenterAM': 'SCM',
    'WshCommonAM': 'SCM', 'WshDeliveriesAM': 'SCM', 'WshCarriersAM': 'SCM',
    'CostElementAM': 'SCM', 'CstPeriodStatusAM': 'SCM',
    'WOMaterialTransactionsAM': 'SCM', 'WOOperationTransactionsAM': 'SCM',
    'WOResourceTransactionsAM': 'SCM', 'WorkOrderCostAM': 'SCM',
    'COGSDistributionsAM': 'SCM', 'CostAnalysisGroupAM': 'SCM',
    'CostBookAM': 'SCM', 'CostTransactionTypeAM': 'SCM',
    'InspectionLevelAM': 'SCM', 'LandedCostChargeAM': 'SCM',
    'ReviewLandedCostChargesAM': 'SCM', 'CustomerAssetAM': 'SCM',
    'ResourceAM': 'SCM', 'ResourceRateAM': 'SCM', 'ResourcesAnalyticsAM': 'SCM',
    # FIN — additional (FA, CE, XLA, VRM, Tax, Interco)
    'FinFaSharedUtilAM': 'FIN', 'FinExmEntrySharedAM': 'FIN',
    'FinFunIntercoTransactionsAM': 'FIN', 'FinCeBankStatementsAM': 'FIN',
    'FinVrmRCSharedPublicModelAM': 'FIN', 'FinVrmShrdSetupPublicModelAM': 'FIN',
    'FinFaDeprDepreciationAM': 'FIN', 'FinFaDeprnSetupCategoriesAM': 'FIN',
    'FinFaTrackDescDetailsAM': 'FIN', 'FinCcControlEnginePublicModelAM': 'FIN',
    'FinCeBankRelationshipsAM': 'FIN', 'FinCeCashTransactionsAM': 'FIN',
    'FinCollSharedAM': 'FIN', 'FinFaSharedSetupBookCtrlsAM': 'FIN',
    'FinLeLegalEntitiesAM': 'FIN', 'FinXlaAmsSuppRefAM': 'FIN',
    'FinXlaSharedObjAM': 'FIN', 'FinVrmRRSharedPublicModelAM': 'FIN',
    'FinFaAssetsSetupLocationAM': 'FIN', 'FinFaRetRetirementsAM': 'FIN',
    'FinFaTrackTransfersAM': 'FIN', 'FinVrmRCSetupDimPublicModelAM': 'FIN',
    'FinXlaAmsEventModelsAM': 'FIN', 'FinXlaAmsSetupSlamAM': 'FIN',
    'FinXlaBalInqSuppRefBalAM': 'FIN', 'FinCcBudgetAccountPublicModelAM': 'FIN',
    'FiscalDocumentAM': 'FIN', 'FiscalDocumentInterfaceAM': 'FIN',
    'AccountBIAM': 'FIN', 'TransactionManagementAM': 'FIN',
    'TradeOperationAM': 'FIN',
    # OKC (Contracts) — additional
    'CnLookupAM': 'OKC', 'BillingEventAM': 'OKC',
    'EligibleCategoryAM': 'OKC',
    # HCM — additional
    'RelationshipsAnalyticsAM': 'HCM', 'PartiesAnalyticsAM': 'HCM',
    'LocationAM': 'HCM', 'JobAM': 'HCM', 'ManagerHierarchyAM': 'HCM',
    'UserAM': 'HCM', 'GeographiesAnalyticsAM': 'HCM',
    # SRP (Sales/Incentive Compensation)
    'SrpCompPlanAM': 'SRP', 'ParticipantSetupAM': 'SRP', 'FormulaAM': 'SRP',
    'RateTableAM': 'SRP', 'CommonAnalyticsAM': 'SRP', 'ScenariosAM': 'SRP',
    'RuleAM': 'SRP', 'ResultsAM': 'SRP', 'ReferenceAM': 'SRP',
    'DisputeAM': 'SRP', 'PartnerCenterAnalyticsAM': 'SRP',
    'PartnerProgramAnalyticsAM': 'SRP', 'RcsCommonPublicViewAM': 'SRP',
    # HED (Higher Education)
    'HedHeyStdExtAcademicInfoAM': 'HED', 'HedHerCurriculumSetupAM': 'HED',
    'HedHesTuitionSetupAM': 'HED', 'HedHesSharedCustAccountAM': 'HED',
    # MFG (Manufacturing / Orchestration)
    'FosOrchestrationProcessAM': 'MFG', 'ServiceRequestAM': 'MFG',
}

EXCEL_FILES = {
    'FSCM': 'src/raw-files/Rel13_25A_BICC_FSCM_Database_Mapping_with_ViewObjects.xlsx',
    'HCM': 'src/raw-files/Rel13_25A_BICC_HCM_Database_Mapping_with_ViewObjects.xlsx',
}


def classify_module(vo_path: str) -> str:
    """Classifica um PVO em módulo baseado no path."""
    parts = vo_path.split('.')
    am_name = parts[1] if len(parts) > 1 else ''

    # 1. Try exact AM name match first
    if am_name in AM_MODULE_MAP:
        return AM_MODULE_MAP[am_name]

    # 2. Try keyword hints on AM name
    for mod, keywords in MODULE_HINTS.items():
        for kw in keywords:
            if kw in am_name:
                return mod

    # 3. Try keyword hints on full path
    for mod, keywords in MODULE_HINTS.items():
        for kw in keywords:
            if kw in vo_path:
                return mod

    return 'OTHER'


def extract_pvos_from_excel(excel_path: str) -> dict:
    """Extrai PVOs de um Excel BICC mapping. Retorna {vo_path: {name, attributes}}."""
    wb = openpyxl.load_workbook(excel_path, read_only=True)

    # Find the mapping sheet
    sheet_name = None
    for name in wb.sheetnames:
        if 'Database' in name or 'BIVO' in name:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[-1]

    ws = wb[sheet_name]
    pvos = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        vo = row[0]
        if not vo or not isinstance(vo, str):
            continue

        if vo not in pvos:
            parts = vo.split('.')
            pvos[vo] = {
                'name': parts[-1],
                'full_path': vo,
                'module_path': '.'.join(parts[:-1]) if len(parts) > 1 else '',
                'attributes': [],
            }

        pvos[vo]['attributes'].append({
            'attribute': row[1] if len(row) > 1 else '',
            'table': row[2] if len(row) > 2 else '',
            'column': row[3] if len(row) > 3 else '',
            'bicc_enabled': (row[4] == 'Yes') if len(row) > 4 else False,
            'otbi_dependent': (row[5] == 'Yes') if len(row) > 5 else False,
            'is_pk': (row[6] == 'Yes') if len(row) > 6 else False,
            'is_incremental': (row[7] == 'Yes') if len(row) > 7 else False,
        })

    wb.close()
    return pvos


def _infer_description(pvo_name: str, tables: list, module: str) -> str:
    """Infere descrição do PVO a partir do nome e tabelas base."""
    name = pvo_name.replace('PVO', '').replace('BPVO', '').replace('ExtractPVO', ' (Extract)')
    # CamelCase to words
    words = re.sub(r'([a-z])([A-Z])', r'\1 \2', name).strip()

    tables_str = ', '.join(tables[:3])
    if len(tables) > 3:
        tables_str += f' (+{len(tables)-3})'

    return f'View Object para extração BICC de dados de {words}. Acessa as tabelas: {tables_str}.'


def build_pvo_json(vo_path: str, data: dict, module: str) -> dict:
    """Constroi o JSON de um PVO para o template HTML."""
    attrs = data['attributes']
    tables = sorted(set(a['table'] for a in attrs if a['table']))
    bicc_count = sum(1 for a in attrs if a['bicc_enabled'])
    pk_count = sum(1 for a in attrs if a['is_pk'])
    bicc_pct = round(bicc_count / len(attrs) * 100) if attrs else 0

    # Group attributes by table
    by_table = {}
    for a in attrs:
        t = a['table'] or 'Unknown'
        if t not in by_table:
            by_table[t] = []
        by_table[t].append(a)

    # Infer description from PVO name + tables
    description = _infer_description(data['name'], tables, module)

    return {
        'meta': {
            'pvo_name': data['name'],
            'full_path': data['full_path'],
            'module_path': data['module_path'],
            'module_code': module,
            'type': 'pvo',
            'description': description,
        },
        'stats': {
            'total_attributes': len(attrs),
            'tables_count': len(tables),
            'bicc_enabled': bicc_count,
            'bicc_pct': bicc_pct,
            'pk_count': pk_count,
        },
        'tables': tables,
        'attributes': [
            {
                'name': a['attribute'],
                'table': a['table'],
                'column': a['column'],
                'bicc': a['bicc_enabled'],
                'pk': a['is_pk'],
                'incremental': a['is_incremental'],
            }
            for a in attrs
        ],
        'by_table': {
            t: [
                {'attribute': a['attribute'], 'column': a['column'], 'bicc': a['bicc_enabled'], 'pk': a['is_pk']}
                for a in rows
            ]
            for t, rows in by_table.items()
        },
    }


def main():
    parser = argparse.ArgumentParser(description='Extract PVO data from BICC Excel')
    parser.add_argument('--module', help='Module code (GL, AP, AR, PO, HCM)')
    parser.add_argument('--all', action='store_true', help='Extract all modules')
    args = parser.parse_args()

    script_dir = Path(__file__).parent
    project_root = script_dir.parent.parent

    # Load all PVOs from Excel files
    all_pvos = {}
    for source, rel_path in EXCEL_FILES.items():
        excel_path = project_root / rel_path
        if excel_path.exists():
            print(f'Loading {source}: {excel_path.name}...')
            pvos = extract_pvos_from_excel(str(excel_path))
            all_pvos.update(pvos)
            print(f'  {len(pvos)} PVOs loaded')

    print(f'Total PVOs: {len(all_pvos)}')

    # Classify by module
    by_module = {}
    for vo, data in all_pvos.items():
        mod = classify_module(vo)
        if mod not in by_module:
            by_module[mod] = {}
        by_module[mod][vo] = data

    for mod in sorted(by_module.keys()):
        print(f'  {mod}: {len(by_module[mod])} PVOs')

    # Filter by requested module
    targets = list(by_module.keys()) if args.all else [args.module.upper()] if args.module else []
    if not targets:
        parser.print_help()
        return

    for mod in targets:
        if mod not in by_module:
            print(f'SKIP: {mod} has no PVOs')
            continue

        output_dir = project_root / 'src' / 'kb' / 'oracle-fusion-data-dictionary' / 'docs' / mod / 'pvos'
        os.makedirs(str(output_dir), exist_ok=True)

        count = 0
        for vo, data in sorted(by_module[mod].items()):
            pvo_json = build_pvo_json(vo, data, mod)
            slug = data['name'].lower()
            output_path = output_dir / f'{slug}.json'

            with open(str(output_path), 'w', encoding='utf-8') as f:
                json.dump(pvo_json, f, ensure_ascii=False, indent=2)
            count += 1

        print(f'{mod}: {count} PVO JSONs saved to {output_dir}')


if __name__ == '__main__':
    main()
