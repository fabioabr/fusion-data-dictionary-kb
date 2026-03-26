#!/usr/bin/env python3
"""
extract_tables.py — Extrai metadados de tabelas do BICC Mapping Excel e gera .md skeleton

Uso:
    python extract_tables.py --module GL
    python extract_tables.py --all
    python extract_tables.py --module GL --force
"""

import argparse
import os
import re
import sys
from datetime import date
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)


# ── Constantes ──────────────────────────────────────────────────────────────

MODULE_META = {
    'GL': 'General Ledger',
    'AP': 'Accounts Payable',
    'AR': 'Accounts Receivable',
    'PO': 'Procurement',
    'HCM': 'Human Capital Management',
}

MODULE_TAGS = {
    'GL': 'general-ledger',
    'AP': 'accounts-payable',
    'AR': 'accounts-receivable',
    'PO': 'procurement',
    'HCM': 'human-capital-management',
}

EXCEL_FILES = {
    'FSCM': 'src/raw-files/Rel13_25A_BICC_FSCM_Database_Mapping_with_ViewObjects.xlsx',
    'HCM': 'src/raw-files/Rel13_25A_BICC_HCM_Database_Mapping_with_ViewObjects.xlsx',
}


# ── Parser de table-mappings.txt ────────────────────────────────────────────

def load_table_map(filepath: str) -> dict:
    """Lê table-mappings.txt e retorna {TABLE_NAME: MODULE_CODE}."""
    table_map = {}
    current_module = None

    with open(filepath, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()

            # Detect module header: === GL — General Ledger (...) ===
            m = re.match(r'^=== (\w+) —', line)
            if m:
                current_module = m.group(1)
                continue

            # Skip non-table lines
            if not line or line.startswith('=') or line.startswith('-') or line.startswith('['):
                continue
            if line.startswith('Fonte:') or line.startswith('Gerado:') or line.startswith('Confiança:'):
                continue
            if line.startswith('REGRAS:') or line.startswith('ORACLE'):
                continue

            # Referenced tables line: [PREFIX] TABLE1, TABLE2, ...
            if line.startswith('['):
                continue

            # Section headers
            if line.startswith('--- '):
                # "--- Core Tables (owned by XX) ---" or "--- Referenced Tables ---"
                if 'Referenced' in line:
                    current_module = None  # stop adding core tables
                continue

            # Core table name (plain line with a valid table name)
            if current_module and re.match(r'^[A-Z][A-Z0-9_]+$', line):
                table_map[line] = current_module

    return table_map


# ── Extração do Excel ───────────────────────────────────────────────────────

def extract_tables_from_excel(excel_path: str) -> dict:
    """
    Extrai tabelas e suas colunas do Excel BICC mapping.

    Retorna {table_name: {columns: {col_name: {pvos: set, is_pk: bool, bicc_enabled: bool}}}}
    """
    wb = openpyxl.load_workbook(excel_path, read_only=True)

    # Find the mapping sheet
    sheet_name = None
    for name in wb.sheetnames:
        if 'Database' in name or 'BIVO' in name:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[-1]

    ws = wb[sheet_name]
    tables = {}

    for row in ws.iter_rows(min_row=2, values_only=True):
        vo = row[0] if len(row) > 0 else None
        table = row[2] if len(row) > 2 else None
        column = row[3] if len(row) > 3 else None

        if not table or not isinstance(table, str):
            continue
        if not column or not isinstance(column, str):
            continue

        bicc_enabled = (row[4] == 'Yes') if len(row) > 4 else False
        is_pk = (row[6] == 'Yes') if len(row) > 6 else False
        pvo_name = vo if vo and isinstance(vo, str) else ''

        if table not in tables:
            tables[table] = {'columns': {}}

        cols = tables[table]['columns']
        if column not in cols:
            cols[column] = {
                'pvos': set(),
                'is_pk': False,
                'bicc_enabled': False,
            }

        if pvo_name:
            cols[column]['pvos'].add(pvo_name)
        if is_pk:
            cols[column]['is_pk'] = True
        if bicc_enabled:
            cols[column]['bicc_enabled'] = True

    wb.close()
    return tables


# ── Geração do .md skeleton ─────────────────────────────────────────────────

def generate_md(table_name: str, columns: dict, module_code: str, seq_num: int, today_str: str) -> str:
    """Gera o conteúdo .md skeleton para uma tabela."""
    module_name = MODULE_META.get(module_code, module_code)
    module_tag = MODULE_TAGS.get(module_code, module_code.lower())
    doc_id = f'DOC-{module_code}-{seq_num:03d}'

    # Build column rows
    col_rows = []
    for i, (col_name, col_info) in enumerate(sorted(columns.items()), start=1):
        col_rows.append(f'| {i} | {col_name} | — | — | — | — | — | — |')

    columns_table = '\n'.join(col_rows)

    md = f"""---
id: {doc_id}
doc_type: system-doc
title: "{table_name} — (título a preencher)"
system: Oracle Fusion Cloud ERP
module: {module_name}
domain: Técnico
owner: fabio.patria
team: dados
status: draft
confidentiality: internal
tags:
  - oracle-fusion
  - {module_tag}
  - data-dictionary
aliases:
  - {table_name}
  - {table_name.lower()}
source_format: markdown
conversion_pipeline: extract_tables-v1
qa_score: 0
qa_status: not_reviewed
created_at: {today_str}
updated_at: {today_str}
---

# {table_name}

## 📌 Visão Geral

(a ser preenchido na etapa 03)

---

## ⚙️ Colunas Principais

| # | Coluna | Tipo | Nulo? | Categoria | Descrição | FK | Confiança |
|---|--------|------|-------|-----------|-----------|-----|-----------|
{columns_table}

---

## 🔗 Relacionamentos

### Tabelas-pai (FK de entrada)

### Tabelas-filha (FK de saída)

---

## 📎 Uso Típico

(a ser preenchido na etapa 03)
"""
    return md


# ── Main ────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Extract table metadata from BICC Excel and generate .md skeletons')
    parser.add_argument('--module', help='Module code (GL, AP, AR, PO, HCM)')
    parser.add_argument('--all', action='store_true', help='Extract all modules')
    parser.add_argument('--force', action='store_true', help='Overwrite existing .md files')
    args = parser.parse_args()

    script_dir = Path(__file__).parent
    project_root = script_dir.parent.parent

    # ── Load table-to-module mapping ──
    mappings_path = project_root / 'src' / 'kb' / 'oracle-fusion-data-dictionary' / 'docs' / 'table-mappings.txt'
    if not mappings_path.exists():
        print(f'ERRO: table-mappings.txt não encontrado em {mappings_path}')
        sys.exit(1)

    table_map = load_table_map(str(mappings_path))
    print(f'Table mappings loaded: {len(table_map)} tabelas mapeadas')

    # ── Load all tables from Excel files ──
    all_tables = {}
    for source, rel_path in EXCEL_FILES.items():
        excel_path = project_root / rel_path
        if excel_path.exists():
            print(f'Loading {source}: {excel_path.name}...')
            tables = extract_tables_from_excel(str(excel_path))
            # Merge tables
            for tbl, data in tables.items():
                if tbl not in all_tables:
                    all_tables[tbl] = data
                else:
                    # Merge columns
                    for col, col_info in data['columns'].items():
                        if col not in all_tables[tbl]['columns']:
                            all_tables[tbl]['columns'][col] = col_info
                        else:
                            existing = all_tables[tbl]['columns'][col]
                            existing['pvos'] |= col_info['pvos']
                            if col_info['is_pk']:
                                existing['is_pk'] = True
                            if col_info['bicc_enabled']:
                                existing['bicc_enabled'] = True
            print(f'  {len(tables)} tabelas extraídas')
        else:
            print(f'WARN: Excel não encontrado: {excel_path}')

    print(f'Total tabelas únicas no Excel: {len(all_tables)}')

    # ── Classify by module using table_map ──
    by_module = {}
    unmapped = 0
    for tbl in all_tables:
        mod = table_map.get(tbl)
        if not mod:
            unmapped += 1
            continue
        if mod not in by_module:
            by_module[mod] = []
        by_module[mod].append(tbl)

    print(f'\nTabelas por módulo (via table-mappings.txt):')
    for mod in sorted(by_module.keys()):
        print(f'  {mod}: {len(by_module[mod])} tabelas')
    if unmapped:
        print(f'  (não mapeadas: {unmapped})')

    # ── Filter by requested module ──
    if args.all:
        targets = [m for m in sorted(by_module.keys()) if m in MODULE_META]
    elif args.module:
        targets = [args.module.upper()]
    else:
        parser.print_help()
        return

    today_str = date.today().isoformat()

    for mod in targets:
        if mod not in by_module:
            print(f'\nSKIP: {mod} não tem tabelas mapeadas')
            continue

        module_name = MODULE_META.get(mod, mod)
        output_dir = project_root / 'src' / 'kb' / 'oracle-fusion-data-dictionary' / 'docs' / mod / 'tables'
        os.makedirs(str(output_dir), exist_ok=True)

        tables_sorted = sorted(by_module[mod])
        created = 0
        skipped = 0

        for seq, tbl in enumerate(tables_sorted, start=1):
            slug = tbl.lower()
            output_path = output_dir / f'{slug}.md'

            if output_path.exists() and not args.force:
                skipped += 1
                continue

            columns = all_tables[tbl]['columns']
            md_content = generate_md(tbl, columns, mod, seq, today_str)

            with open(str(output_path), 'w', encoding='utf-8') as f:
                f.write(md_content)
            created += 1

        print(f'\n{mod} ({module_name}):')
        print(f'  Criados: {created} arquivos .md')
        if skipped:
            print(f'  Ignorados (já existem): {skipped} — use --force para sobrescrever')
        print(f'  Diretório: {output_dir}')


if __name__ == '__main__':
    main()
